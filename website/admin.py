from django.contrib import admin,messages
from .models import AdminFeautures, BonusEmailsHistory, Employee, EvaluationFormModel, Teams
from django import forms
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from datetime import date
from collections import defaultdict
import openpyxl
from django.http import HttpResponse

# Register your models here.

admin.site.site_title = "Highbit HR Admin Panel"
admin.site.site_header = "HR Admin Panel"


class TeamFilter(admin.SimpleListFilter):
    title = "Team"
    parameter_name = "team"
    
    def lookups(self, request, model_admin):
        teams = Teams.objects.all()
        return [(team.id, team.team_name) for team in teams]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(employee__team_id=self.value())
        return queryset


def send_evaluation_email(modeladmin, request, queryset):
    email_count=0
    for evaluation in queryset:
        employee = evaluation.employee
        team_lead_name = (
            employee.team_lead.employee_name if employee.team_lead else "N/A"
        )
        email_template = "evaluation_email.html"
        context = {
            "employee_name": employee.employee_name,
            "team_lead_name": team_lead_name,
            "evaluation_for": evaluation.evaluation_for,
            "hr_marks": evaluation.hr_marks,
            "tl_marks": evaluation.tl_marks,
            "weightage": evaluation._weighted_average,
            "feedback": evaluation.feedback,
        }
        html_message = render_to_string(email_template, context)
        plain_message = strip_tags(html_message)
        performance_review = f"Performance Review for {evaluation.evaluation_for}"
        send_mail(
            performance_review,
            plain_message,
            "your_email@example.com",
            [employee.employee_email],
            html_message=html_message,
        )

        email_count += 1

    messages.success(request, f"Successfully sent {email_count} monthly evaluation emails.")
send_evaluation_email.short_description = (
    "Send monthly evaluation email to selected employees"
)


class IsTeamLeadFilter(admin.SimpleListFilter):
    title = "Team Lead"
    parameter_name = "team_lead"

    def lookups(self, request, model_admin):
        team_lead_names = Employee.objects.filter(is_team_lead=True).values_list(
            "employee_name", flat=True
        )
        return [(name, name) for name in team_lead_names]

    def queryset(self, request, queryset1):
        if self.value():
            print(self.value())
            return Employee.objects.filter(team_lead__employee_name=self.value())
        else:
            return queryset1


class EmployeeAdminForm(forms.ModelForm):
    class Meta:
        model = Employee
        exclude = ["username"]


def calculate_quarterly_average(employee, end_date):
    start_month = end_date.month - 3
    start_year = end_date.year if start_month > 0 else end_date.year - 1
    start_month = start_month % 12 if start_month > 0 else start_month + 12

    target_months = []
    for i in range(3):
        target_month = (start_month + i - 1) % 12 + 1
        target_year = start_year if start_month + i <= 12 else start_year + 1
        month_name = date(1900, target_month, 1).strftime("%B")
        target_months.append(f"{month_name} {target_year}")

    evaluations = EvaluationFormModel.objects.filter(
        employee=employee, evaluation_for__in=target_months
    )

    if not evaluations.exists():
        return None, None, None

    monthly_ratings = defaultdict(list)
    for evaluation in evaluations:
        monthly_ratings[evaluation.evaluation_for].append(evaluation._weighted_average)

    month_averages = {
        month: sum(scores) / len(scores) for month, scores in monthly_ratings.items()
    }

    average_weighted_score = sum(month_averages.values()) / len(month_averages)

    if average_weighted_score >= 8.5:
        grade = "A"
    elif average_weighted_score >= 7.5:
        grade = "B"
    else:
        grade = "C"

    return average_weighted_score, grade, month_averages


def send_quarterly_evaluation_email(modeladmin, request, queryset):
    end_date = date.today()
    email_count = 0

    for employee in queryset:
        average_weighted_score, grade, month_averages = calculate_quarterly_average(
            employee, end_date
        )

        if average_weighted_score is None:
            continue

        team_lead_name = (
            employee.team_lead.employee_name if employee.team_lead else "N/A"
        )
        team_name = employee.team.team_name if employee.team else "N/A"

        email_template = "quarterly_evaluation_email.html"
        context = {
            "employee_name": employee.employee_name,
            "team_lead_name": team_lead_name,
            "team_name": team_name,
            "average_weighted_score": average_weighted_score,
            "grade": grade,
            "month_averages": month_averages,
        }

        html_message = render_to_string(email_template, context)
        plain_message = strip_tags(html_message)
        performance_review = f"Quarterly Performance Review - Q{(end_date.month - 2) // 3 + 1} {end_date.year}"

        send_mail(
            performance_review,
            plain_message,
            "your_email@example.com",
            [employee.employee_email],
            html_message=html_message,
        )
        email_count += 1

    messages.success(request, f"Successfully sent {email_count} quarterly evaluation emails.")

send_quarterly_evaluation_email.short_description = (
    "Send quarterly evaluation email to selected employees"
)


class EmployeeInformation(admin.ModelAdmin):
    list_display = ("employee_id", "employee_name","role","is_permanent","is_active", "team", "team_lead")
    readonly_fields = ("is_permanent",)
    search_fields = ["employee_id", "employee_name"]
    list_filter = ["team", "is_team_lead", IsTeamLeadFilter]
    actions = [send_quarterly_evaluation_email]
    exclude = ["username", "password", "grade"]
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "team_lead":
            kwargs["queryset"] = Employee.objects.filter(is_team_lead=True)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


class WeightedAverageFilter(admin.SimpleListFilter):
    title = "Weighted Average Calculated"
    parameter_name = "_weighted_average"

    def lookups(self, request, model_admin):
        return (
            ("yes", "Yes"),
            ("no", "No"),
        )

    def queryset(self, request, queryset):
        if self.value() == "yes":
            return queryset.exclude(_weighted_average__isnull=True)
        if self.value() == "no":
            return queryset.filter(_weighted_average__isnull=True)
        return queryset


def export_to_excel(modeladmin, request, queryset):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=evaluation_form.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EvaluationForm"

    columns = [
        "Employee",
        "Evaluated By",
        "TL Marks",
        "HR Marks",
        "Feedback",
        "Weighted Average",
        "Evaluation Date",
        "Evaluation For",
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = column_title

    for obj in queryset:
        row_num += 1
        row = [
            obj.employee.employee_name if obj.employee else "",
            (
                obj.employee.team_lead.employee_name
                if obj.employee and obj.employee.team_lead
                else ""
            ),
            obj.tl_marks,
            obj.hr_marks,
            obj.feedback,
            obj._weighted_average,
            obj.evaluation_date,
            obj.evaluation_for,
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
    wb.save(response)
    return response


export_to_excel.short_description = "Export Selected to Excel"


class EvaluationFormModelAdmin(admin.ModelAdmin):
    list_display = [
        "employee_id",
        "employee",
        "employee_role",
        "employee_email",
        "employee_team",
        "evaluated_by",
        "evaluation_for",
        "evaluation_date",
        "_weighted_average",  
    ]
    list_filter = ["evaluation_for", TeamFilter, WeightedAverageFilter]
    search_fields = ["employee__employee_name", "employee__employee_id"]
    actions = [send_evaluation_email, export_to_excel]

    def get_weighted_average(self, obj):
        return obj._weighted_average

    def employee_id(self, obj):
        return obj.employee.employee_id

    def employee_email(self, obj):
        return obj.employee.employee_email

    def employee_team(self, obj):
        return obj.employee.team
    
    def employee_role(self, obj):
        return obj.employee.role

    get_weighted_average.short_description = "Weighted Average"
    get_weighted_average.admin_order_field = '_weighted_average'  


class TeamsAdmin(admin.ModelAdmin):
    list_display = ("team_name", "evaluations_submitted_status")
    list_filter = ("team_name",)

    def evaluations_submitted_status(self, obj):
        return obj.evaluations_submitted

    evaluations_submitted_status.boolean = True
    evaluations_submitted_status.short_description = "Evaluations Submitted"


admin.site.register(Teams, TeamsAdmin)
admin.site.register(Employee, EmployeeInformation)
admin.site.register(EvaluationFormModel, EvaluationFormModelAdmin)
admin.site.register(AdminFeautures)
admin.site.register(BonusEmailsHistory)